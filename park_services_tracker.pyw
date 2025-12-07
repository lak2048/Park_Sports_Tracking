# -*- coding: utf-8 -*-
# park_services_tracker.pyw
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import date
import pandas as pd
import os
from tkcalendar import DateEntry
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import time

class ParkServicesTrackerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Park Services Tracker - Sports & Non-Sports")
        self.root.geometry("900x700")
        
        # File path for auto-save (Excel only)
        self.excel_file = "park_services_dashboard.xlsx"
        
        self.all_data = []
        
        # Updated fee structure with new services
        self.fee_structure = {
            'Railway': {
                # Sports Services
                'Cricket': 800, 'Football': 500, 'Athletic': 300, 'KGS': 400,
                'Silambam-M': 500, 'Silambam-E': 500,
                # Non-Sports Services (same price for all categories)
                'Gym': 300, 'Walking(1+1)': 400, 'Walking(6 month)': 1200, 
                'Car': 200, 'Walking': 200, 'Bike': 50
            },
            'Non Railway': {
                # Sports Services
                'Cricket': 1200, 'Football': 700, 'Athletic': 500, 'KGS': 600,
                'Shuttle': 1500, 'Volleyball': 3000, 'Team Football': 3000,
                'Silambam-M': 800, 'Silambam-E': 800,
                # Non-Sports Services (same price for all categories)
                'Gym': 300, 'Walking(1+1)': 400, 'Walking(6 month)': 1200,
                'Car': 200, 'Walking': 200, 'Bike': 50
            }
        }
        
        # Updated service categories
        self.service_categories = {
            'Sports': ['Cricket', 'Football', 'Athletic', 'Shuttle', 'Volleyball', 
                      'Team Football', 'Silambam-M', 'Silambam-E', 'KGS'],
            'Non-Sports': ['Gym', 'Walking(1+1)', 'Walking(6 month)', 'Car', 'Walking', 'Bike']
        }
        
        # Services that default to Non Railway and disable category selection
        self.non_railway_only_services = ['Shuttle', 'Volleyball', 'Team Football']
        
        # Initialize report_tree as None
        self.report_tree = None
        
        # *** IMPORTANT: Load existing data when program starts ***
        self.load_existing_data()
        
        # Clean any duplicate data that might exist
        self.clean_duplicate_data()
        
        self.notebook = ttk.Notebook(root)
        self.notebook.pack(pady=10, padx=10, fill="both", expand=True)
        
        self.data_entry_frame = ttk.Frame(self.notebook)
        self.admin_dashboard_frame = ttk.Frame(self.notebook)
        
        self.notebook.add(self.data_entry_frame, text="Service Entry")
        self.notebook.add(self.admin_dashboard_frame, text="Admin Dashboard")
        
        self.create_data_entry_tab()
        self.create_admin_dashboard_tab()
        
        # Load existing data into the report view immediately (AFTER report_tree is created)
        self.generate_report()
        
        # Initialize status variable
        self.status_var = tk.StringVar()
    
    def get_service_category(self, service):
        """Get service category based on service name"""
        for category, services in self.service_categories.items():
            if service in services:
                return category
        return 'Sports'  # Default fallback
    
    def get_next_bill_no(self):
        """Generate next bill number from 0001 to 2000, then reset to 0001"""
        if not self.all_data:
            return "0001"
        
        # Find the highest bill number
        bill_numbers = []
        for record in self.all_data:
            try:
                bill_num = int(record['Token No'])
                bill_numbers.append(bill_num)
            except:
                pass
        
        if not bill_numbers:
            return "0001"
        
        max_bill = max(bill_numbers)
        next_bill = max_bill + 1
        
        # Reset to 0001 after 2000
        if next_bill > 2000:
            next_bill = 1
        
        return f"{next_bill:04d}"
    
    def check_bill_no_exists(self, bill_no):
        """Check if token number already exists"""
        for record in self.all_data:
            if record['Token No'] == bill_no:
                return True
        return False
    
    def migrate_old_data(self, records):
        """Migrate old data structure to new format with Service Category"""
        migrated_records = []
        
        for record in records:
            # If record doesn't have 'Service Category', add it
            if 'Service Category' not in record:
                # Check if it's old format (has 'Sport' instead of 'Service')
                if 'Sport' in record:
                    service = record['Sport']
                    record['Service'] = service
                    del record['Sport']  # Remove old key
                elif 'Service' not in record:
                    # Very old format, set default
                    record['Service'] = 'Cricket'
                
                # Add Service Category based on Service
                service = record.get('Service', 'Cricket')
                
                # Handle old naming migrations
                if service == 'Car Parking':
                    record['Service'] = 'Car'
                elif service == 'Bike Parking':
                    record['Service'] = 'Bike'
                elif service == 'KG':
                    record['Service'] = 'KGS'
                elif service == 'Silambam':
                    record['Service'] = 'Silambam-M'  # Default to Silambam-M
                elif service == 'Valleyball':  # Migration for old Valleyball records
                    record['Service'] = 'Volleyball'
                
                service = record['Service']  # Update service after migration
                record['Service Category'] = self.get_service_category(service)
                
                print(f"Migrated record: {service} -> {record['Service Category']} - {record['Service']}")
            
            migrated_records.append(record)
        
        return migrated_records
    
    def load_existing_data(self):
        """Load data from Excel file if it exists with migration support"""
        try:
            if os.path.exists(self.excel_file):
                # Read Excel file
                df = pd.read_excel(self.excel_file, dtype={'Token No': str})
                
                # Remove any rows where 'Name' is 'Total' or empty (these are total rows)
                df = df[~df['Name'].isin(['Total', 'nan', ''])]
                df = df.dropna(subset=['Name'])
                
                # Ensure Token No is properly formatted with leading zeros
                df['Token No'] = df['Token No'].apply(lambda x: f"{int(x):04d}" if str(x).isdigit() else str(x))
                
                # Remove any duplicates that might exist in the file
                df = df.drop_duplicates(subset=['Token No'], keep='last')
                
                # Convert to dict and migrate old data structure
                records = df.to_dict('records')
                migrated_records = self.migrate_old_data(records)
                
                self.all_data = migrated_records
                
                print(f"Loaded and migrated {len(self.all_data)} existing records from {self.excel_file}")
            else:
                self.all_data = []
                print("No existing Excel file found. Starting fresh.")
        except Exception as e:
            print(f"Error loading data: {e}")
            self.all_data = []
    
    def clean_duplicate_data(self):
        """Clean any existing duplicate data"""
        if self.all_data:
            df = pd.DataFrame(self.all_data)
            original_count = len(df)
            df = df.drop_duplicates(subset=['Token No'], keep='last')
            self.all_data = df.to_dict('records')
            
            if len(df) < original_count:
                self.save_data_automatically()
                print(f"Cleaned {original_count - len(df)} duplicate records")
                return True
        return False
    
    def save_data_automatically(self):
        """Automatically save data to Excel file with professional styling and file locking handling"""
        max_retries = 3
        retry_delay = 1  # seconds
        
        for attempt in range(max_retries):
            try:
                if self.all_data:
                    df = pd.DataFrame(self.all_data)
                    df['Token No'] = df['Token No'].astype(str)
                    
                    # Ensure all records have the required columns
                    required_columns = ['Token No', 'Date', 'Name', 'Service Category', 'Service', 'Category', 'Fee']
                    for col in required_columns:
                        if col not in df.columns:
                            if col == 'Service Category':
                                df['Service Category'] = df['Service'].apply(self.get_service_category)
                            else:
                                df[col] = ''
                    
                    # Reorder columns
                    df = df[required_columns]
                    
                    # Remove duplicates based on Token No before saving
                    df = df.drop_duplicates(subset=['Token No'], keep='last')
                    
                    # Try to delete existing file with retry logic
                    if os.path.exists(self.excel_file):
                        try:
                            os.remove(self.excel_file)
                        except PermissionError:
                            if attempt < max_retries - 1:
                                print(f"File in use, retrying in {retry_delay} seconds... (attempt {attempt + 1})")
                                time.sleep(retry_delay)
                                continue
                            else:
                                print("Warning: Could not delete existing file, will try to overwrite")
                    
                    # Save to Excel
                    with pd.ExcelWriter(self.excel_file, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False, sheet_name='Park Services Report')
                        self.style_excel_professional(writer.book.active, df)
                    
                    # Update self.all_data to match what was saved
                    self.all_data = df.to_dict('records')
                    
                    print(f"Data automatically saved to {self.excel_file} with {len(df)} records")
                    break  # Success, exit retry loop
                    
            except Exception as e:
                if attempt < max_retries - 1:
                    print(f"Error saving data (attempt {attempt + 1}): {e}. Retrying...")
                    time.sleep(retry_delay)
                else:
                    print(f"Final error saving data: {e}")
    
    def style_excel_professional(self, worksheet, df):
        """Apply professional styling with service-based colors and Railway bold formatting - SAME SIZE"""
        # Clear any existing content first (including old total rows)
        max_row = worksheet.max_row
        if max_row > len(df) + 1:  # If there are more rows than data + header
            for row in range(len(df) + 2, max_row + 1):
                for col in range(1, len(df.columns) + 1):
                    worksheet.cell(row=row, column=col).value = None
        
        # Define colors and styles
        header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow header
        sports_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')  # Light Blue for sports
        non_sports_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Light Green for non-sports
        total_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow for total
        
        # Font and alignment - FIXED: Same size for bold and regular text
        header_font = Font(bold=True, color='000000', size=11)
        data_font = Font(size=10, color='000000', bold=False)  # Regular font
        railway_font = Font(size=10, color='000000', bold=True)  # Bold font - SAME SIZE as regular
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Thin border for all cells
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Style header row
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
            
            # Set column widths for better appearance
            column_letter = get_column_letter(col_num)
            if col_num == 1:  # Token No
                worksheet.column_dimensions[column_letter].width = 10
            elif col_num == 2:  # Date
                worksheet.column_dimensions[column_letter].width = 12
            elif col_num == 3:  # Name
                worksheet.column_dimensions[column_letter].width = 18
            elif col_num == 4:  # Service Category
                worksheet.column_dimensions[column_letter].width = 15
            elif col_num == 5:  # Service
                worksheet.column_dimensions[column_letter].width = 15
            elif col_num == 6:  # Category
                worksheet.column_dimensions[column_letter].width = 12
            elif col_num == 7:  # Fee
                worksheet.column_dimensions[column_letter].width = 10
        
        # Style data rows with service category-based colors
        for row_num, (_, row_data) in enumerate(df.iterrows(), start=2):
            service_category = row_data.get('Service Category', 'Sports')
            category = row_data.get('Category', '')
            
            # Choose color based on service category
            if service_category == 'Sports':
                row_fill = sports_fill  # Blue for sports
            else:  # Non-Sports
                row_fill = non_sports_fill  # Green for non-sports
            
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.fill = row_fill
                cell.alignment = center_alignment
                cell.border = thin_border
                
                # FIXED: Bold formatting for Railway categories in Sports - SAME SIZE
                if service_category == 'Sports' and category == 'Railway':
                    cell.font = railway_font  # Bold font, same size
                else:
                    cell.font = data_font  # Regular font
                
                # Format Token No column as text
                if col_num == 1:
                    cell.number_format = '@'
        
        # Add total row
        total_row = len(df) + 2
        
        # Clear all columns in the total row first
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=total_row, column=col_num)
            cell.value = None
            cell.fill = PatternFill()
            cell.font = Font()
            cell.border = Border()
        
        # Add new total row
        # Empty cells for columns 1-6
        for col_num in range(1, 7):
            cell = worksheet.cell(row=total_row, column=col_num)
            cell.value = ""
            cell.fill = total_fill
            cell.border = thin_border
        
        # "Total" label in column 6
        total_label_cell = worksheet.cell(row=total_row, column=6)
        total_label_cell.value = "Total"
        total_label_cell.fill = total_fill
        total_label_cell.font = Font(bold=True, size=11)
        total_label_cell.alignment = center_alignment
        total_label_cell.border = thin_border
        
        # Calculate and add total amount in column 7
        total_amount = df['Fee'].sum()
        total_amount_cell = worksheet.cell(row=total_row, column=7)
        total_amount_cell.value = total_amount
        total_amount_cell.fill = total_fill
        total_amount_cell.font = Font(bold=True, size=11)
        total_amount_cell.alignment = center_alignment
        total_amount_cell.border = thin_border
        
        # Set row height for better appearance
        worksheet.row_dimensions[1].height = 25  # Header row
        for row_num in range(2, len(df) + 3):  # Data rows + total row
            worksheet.row_dimensions[row_num].height = 20
    
    def dataframe_to_pdf(self, df, filename, title="Park Services Report"):
        """Convert DataFrame to PDF with total row and professional formatting with Railway bold text - SAME SIZE"""
        try:
            # Calculate total and add total row to dataframe
            total_fee = df['Fee'].sum()
            
            # Create a copy of the dataframe and add total row
            df_with_total = df.copy()
            total_row = pd.DataFrame({
                'Token No': [''],
                'Date': [''],
                'Name': ['Total'],
                'Service Category': [''],
                'Service': [''],
                'Category': [''],
                'Fee': [total_fee]
            })
            df_with_total = pd.concat([df_with_total, total_row], ignore_index=True)
            
            # Create figure with appropriate size
            fig, ax = plt.subplots(figsize=(16, max(6, 0.28*len(df_with_total))))
            ax.axis('off')
            
            # Add title with reduced margin
            fig.suptitle(title, fontsize=16, fontweight='bold', y=0.98)
            
            # Create table
            table = ax.table(cellText=df_with_total.values,
                           colLabels=df_with_total.columns,
                           loc='center',
                           cellLoc='center')
            
            # Format table - FIXED: Same size for all text
            table.auto_set_font_size(False)
            table.set_fontsize(9)  # Same font size for all
            table.scale(1, 1.5)
            
            # Style header row (Yellow background, black bold text)
            for i in range(len(df_with_total.columns)):
                table[(0, i)].set_facecolor('#FFFF00')  # Yellow header
                table[(0, i)].set_text_props(weight='bold', color='black')
            
            # Style data rows with service category-based colors
            for i in range(1, len(df_with_total) + 1):
                if i == len(df_with_total):  # This is the total row
                    # Style total row (Yellow background, bold text)
                    for j in range(len(df_with_total.columns)):
                        table[(i, j)].set_facecolor('#FFFF00')  # Yellow background
                        table[(i, j)].set_text_props(weight='bold', color='black')
                else:
                    # Regular data rows with service category colors
                    service_category = df_with_total.iloc[i-1].get('Service Category', 'Sports')
                    category = df_with_total.iloc[i-1].get('Category', '')
                    
                    if service_category == 'Sports':
                        row_color = '#ADD8E6'  # Blue for sports
                    else:  # Non-Sports
                        row_color = '#90EE90'  # Green for non-sports
                    
                    for j in range(len(df_with_total.columns)):
                        table[(i, j)].set_facecolor(row_color)
                        
                        # FIXED: Bold text for Railway categories in Sports - SAME SIZE
                        if service_category == 'Sports' and category == 'Railway':
                            table[(i, j)].set_text_props(color='black', weight='bold')  # Bold, same size
                        else:
                            table[(i, j)].set_text_props(color='black', weight='normal')  # Regular, same size
            
            # Save to PDF
            pp = PdfPages(filename)
            pp.savefig(fig, bbox_inches='tight', pad_inches=0.1)
            pp.close()
            plt.close()
            
            return True
        except Exception as e:
            print(f"Error creating PDF: {e}")
            return False
    
    def create_data_entry_tab(self):
        form_frame = ttk.LabelFrame(self.data_entry_frame, text="Park Services Entry", padding="20")
        form_frame.pack(padx=20, pady=20)
        
        # Token No with Edit button
        ttk.Label(form_frame, text="Token No:").grid(row=0, column=0, padx=10, pady=10, sticky="w")
        
        # Frame for Token No display and edit button
        bill_frame = ttk.Frame(form_frame)
        bill_frame.grid(row=0, column=1, padx=10, pady=10, sticky="w")
        
        # Token No display (initially read-only)
        self.bill_no_var = tk.StringVar()
        self.bill_no_display = ttk.Label(bill_frame, textvariable=self.bill_no_var, 
                                        font=("Arial", 10, "bold"), foreground="blue",
                                        relief="sunken", padding=5)
        self.bill_no_display.grid(row=0, column=0, padx=2)
        
        # Token No entry (hidden initially)
        self.bill_no_entry = ttk.Entry(bill_frame, width=10, font=("Arial", 10, "bold"))
        
        # Edit button
        self.edit_button = ttk.Button(bill_frame, text="Edit", command=self.toggle_bill_edit, width=6)
        self.edit_button.grid(row=0, column=1, padx=5)
        
        # Status label for token number
        self.bill_status_var = tk.StringVar()
        self.bill_status_label = ttk.Label(bill_frame, textvariable=self.bill_status_var, font=("Arial", 9))
        self.bill_status_label.grid(row=1, column=0, columnspan=2, pady=2)
        
        # Set initial token number
        self.bill_no_var.set(self.get_next_bill_no())
        self.is_editing_bill = False
        
        # Date with Calendar
        ttk.Label(form_frame, text="Date:").grid(row=1, column=0, padx=10, pady=10, sticky="w")
        self.date_entry = DateEntry(form_frame, width=27, background='darkblue',
                                   foreground='white', borderwidth=2, 
                                   date_pattern='dd-mm-yyyy')
        self.date_entry.grid(row=1, column=1, padx=10, pady=10)
        
        # Name (CHANGED: removed "Vehicle No" reference)
        ttk.Label(form_frame, text="Name:").grid(row=2, column=0, padx=10, pady=10, sticky="w")
        self.name_entry = ttk.Entry(form_frame, width=30)
        self.name_entry.grid(row=2, column=1, padx=10, pady=10)
        
        # Service Category (Main Category Selection)
        ttk.Label(form_frame, text="Service Category:").grid(row=3, column=0, padx=10, pady=10, sticky="w")
        self.service_category_var = tk.StringVar()
        category_options = ['Sports', 'Non-Sports']
        self.service_category_combobox = ttk.Combobox(form_frame, textvariable=self.service_category_var, 
                                                     values=category_options, state='readonly', width=27)
        self.service_category_combobox.set(category_options[0])
        self.service_category_combobox.grid(row=3, column=1, padx=10, pady=10)
        self.service_category_combobox.bind('<<ComboboxSelected>>', self.update_service_dropdown)
        
        # Service (Sub-service Selection)
        ttk.Label(form_frame, text="Service:").grid(row=4, column=0, padx=10, pady=10, sticky="w")
        self.service_var_entry = tk.StringVar()
        self.service_combobox = ttk.Combobox(form_frame, textvariable=self.service_var_entry, 
                                           state='readonly', width=27)
        self.service_combobox.grid(row=4, column=1, padx=10, pady=10)
        self.service_combobox.bind('<<ComboboxSelected>>', self.on_service_entry_change)
        
        # Category (Railway/Non Railway)
        ttk.Label(form_frame, text="Category:").grid(row=5, column=0, padx=10, pady=10, sticky="w")
        self.category_var_entry = tk.StringVar()
        railway_category_options = ['Railway', 'Non Railway']
        self.category_combobox = ttk.Combobox(form_frame, textvariable=self.category_var_entry, 
                                            values=railway_category_options, state='readonly', width=27)
        self.category_combobox.set(railway_category_options[0])
        self.category_combobox.grid(row=5, column=1, padx=10, pady=10)
        self.category_combobox.bind('<<ComboboxSelected>>', self.update_fee_display)
        
        # Calculated Fee
        ttk.Label(form_frame, text="Fee:").grid(row=6, column=0, padx=10, pady=10, sticky="w")
        self.fee_display_var = tk.StringVar()
        fee_label = ttk.Label(form_frame, textvariable=self.fee_display_var, font=("Arial", 10, "bold"))
        fee_label.grid(row=6, column=1, padx=10, pady=10, sticky="w")
        
        # Initialize service dropdown and fee display
        self.update_service_dropdown()
        
        submit_button = ttk.Button(form_frame, text="Issue Token", command=self.submit_data)
        submit_button.grid(row=7, column=0, columnspan=2, pady=20)
        
        # Add status label for feedback
        self.status_var = tk.StringVar()
        self.status_label = ttk.Label(form_frame, textvariable=self.status_var, 
                                     font=("Arial", 10), foreground="green")
        self.status_label.grid(row=8, column=0, columnspan=2, pady=5)
    
    def update_service_dropdown(self, *args):
        """Update service dropdown based on selected service category"""
        selected_category = self.service_category_var.get()
        
        if selected_category in self.service_categories:
            # Get services for selected category
            services = self.service_categories[selected_category]
            
            # Update service combobox
            self.service_combobox['values'] = services
            self.service_combobox.set(services[0])  # Set first service as default
            
            # Handle category auto-selection based on service category
            if selected_category == 'Non-Sports':
                # Non-Sports: Always set to Non Railway and disable
                self.category_var_entry.set('Non Railway')
                self.category_combobox.config(state='disabled')
            else:
                # Sports: Enable category selection initially
                self.category_combobox.config(state='readonly')
                # Check if the first service needs special handling
                self.on_service_entry_change()
            
            # Update fee display
            self.update_fee_display()
    
    def on_service_entry_change(self, *args):
        """Handle service selection changes for category auto-selection"""
        service = self.service_var_entry.get()
        service_category = self.service_category_var.get()
        
        if service_category == 'Sports':
            # For specific sports services, default to Non Railway and disable
            if service in self.non_railway_only_services:
                self.category_var_entry.set('Non Railway')
                self.category_combobox.config(state='disabled')
            else:
                # For other sports, enable category selection
                self.category_combobox.config(state='readonly')
        elif service_category == 'Non-Sports':
            # Non-Sports always Non Railway
            self.category_var_entry.set('Non Railway')
            self.category_combobox.config(state='disabled')
        
        # Update fee display
        self.update_fee_display()
    
    def update_service_filter_dropdown(self, *args):
        """Update service filter dropdown based on selected service category"""
        selected_category = self.service_category_filter_var.get()
        
        if selected_category == 'All':
            # Show all services
            all_services = ['All'] + self.service_categories['Sports'] + self.service_categories['Non-Sports']
        elif selected_category == 'Sports':
            # Show only sports services
            all_services = ['All'] + self.service_categories['Sports']
        elif selected_category == 'Non-Sports':
            # Show only non-sports services  
            all_services = ['All'] + self.service_categories['Non-Sports']
        else:
            all_services = ['All']
        
        # Update service filter combobox
        if hasattr(self, 'service_filter_menu'):
            self.service_filter_menu['values'] = all_services
            self.service_filter_menu.set('All')  # Reset to 'All' when category changes
        
        # Generate report with updated filter (only if report_tree exists)
        if self.report_tree is not None:
            self.generate_report()
    
    def clear_all_filters(self):
        """NEW: Clear all filters and set everything to 'All' - ADDED FUNCTION"""
        # Set date filter to All
        self.date_filter_var.set('All')
        self.date_calendar.configure(state='disabled')
        
        # Set service category filter to All
        self.service_category_filter_var.set('All')
        
        # Set service filter to All
        self.service_var_filter.set('All')
        
        # Set category filter to All
        self.category_var_filter.set('All')
        
        # Update service filter dropdown to show all services
        self.update_service_filter_dropdown()
        
        # Generate report with all filters cleared
        if self.report_tree is not None:
            self.generate_report()
    
    def toggle_bill_edit(self):
        """Toggle between edit and view mode for token number"""
        if not self.is_editing_bill:
            # Switch to edit mode
            current_bill = self.bill_no_var.get()
            
            # Hide label, show entry
            self.bill_no_display.grid_remove()
            self.bill_no_entry.grid(row=0, column=0, padx=2)
            self.bill_no_entry.delete(0, 'end')
            self.bill_no_entry.insert(0, current_bill)
            self.bill_no_entry.focus()
            
            # Change button text
            self.edit_button.configure(text="Save")
            self.is_editing_bill = True
            self.bill_status_var.set("Edit mode - Enter token number and click Save")
            self.bill_status_label.configure(foreground="orange")
            
        else:
            # Switch to view mode
            entered_bill = self.bill_no_entry.get().strip()
            
            if not entered_bill:
                messagebox.showerror("Error", "Token number cannot be empty!")
                return
            
            # Validate and format token number
            try:
                bill_num = int(entered_bill)
                if bill_num < 1 or bill_num > 2000:
                    messagebox.showerror("Error", "Token number must be between 0001-2000!")
                    return
                
                formatted_bill = f"{bill_num:04d}"
                
                # Check if token already exists (only if it's different from current)
                if formatted_bill != self.bill_no_var.get() and self.check_bill_no_exists(formatted_bill):
                    messagebox.showerror("Error", f"Token No {formatted_bill} already exists!")
                    return
                
                # Update the token number
                self.bill_no_var.set(formatted_bill)
                
                # Hide entry, show label
                self.bill_no_entry.grid_remove()
                self.bill_no_display.grid(row=0, column=0, padx=2)
                
                # Change button text
                self.edit_button.configure(text="Edit")
                self.is_editing_bill = False
                
                # Update status
                if formatted_bill != entered_bill:
                    self.bill_status_var.set(f"✅ Token number updated to {formatted_bill}")
                else:
                    self.bill_status_var.set("✅ Token number confirmed")
                self.bill_status_label.configure(foreground="green")
                
            except ValueError:
                messagebox.showerror("Error", "Invalid token number format!")
                return
    
    def update_fee_display(self, *args):
        service = self.service_var_entry.get()
        category = self.category_var_entry.get()
        if service and category:
            fee = self.fee_structure.get(category, {}).get(service, "")
            if fee != "":
                self.fee_display_var.set(f"Rs. {fee}")
            else:
                # Fallback to Non Railway for new sports if odd case
                nr_fee = self.fee_structure['Non Railway'].get(service)
                if nr_fee:
                    self.fee_display_var.set(f"Rs. {nr_fee}")
    
    def submit_data(self):
        # If still in edit mode, force save the token number first
        if self.is_editing_bill:
            self.toggle_bill_edit()
            if self.is_editing_bill:  # If save failed, don't continue
                return
        
        token_no = self.bill_no_var.get()
        reg_date = self.date_entry.get()  # This will return DD-MM-YYYY format
        name = self.name_entry.get()
        service_category = self.service_category_var.get()
        service = self.service_var_entry.get()
        category = self.category_var_entry.get()
        
        if not name.strip():
            messagebox.showerror("Input Error", "Name cannot be empty.")
            return
        
        # Check if this exact token number already exists
        if self.check_bill_no_exists(token_no):
            response = messagebox.askyesno("Duplicate Token Number", 
                                         f"Token No {token_no} already exists!\n\nDo you want to update the existing record?")
            if response:
                # Remove existing record with same token number
                self.all_data = [record for record in self.all_data if record['Token No'] != token_no]
            else:
                return
        
        fee = self.fee_structure[category][service]
        new_record = {
            'Token No': token_no, 
            'Date': reg_date, 
            'Name': name, 
            'Service Category': service_category,
            'Service': service, 
            'Category': category, 
            'Fee': fee
        }
        self.all_data.append(new_record)
        
        # Save data immediately after adding
        self.save_data_automatically()
        
        # Status message instead of popup
        self.status_var.set(f"✅ {service_category} - {service} token issued to {name} (Token: {token_no}) - Rs. {fee}")
        
        # Auto-clear status after 3 seconds
        self.root.after(3000, lambda: self.status_var.set(""))
        
        # Reset form and generate next token number
        self.name_entry.delete(0, 'end')
        self.bill_no_var.set(self.get_next_bill_no())
        self.bill_status_var.set("")
        
        # Update report only if report_tree exists
        if self.report_tree is not None:
            self.generate_report()
        
        # Set focus back to name entry for quick next entry
        self.name_entry.focus()
    
    def on_date_filter_change(self):
        """Called when date filter selection changes"""
        if self.date_filter_var.get() == 'All':
            self.date_calendar.configure(state='disabled')
        else:
            self.date_calendar.configure(state='normal')
        
        # Generate report only if report_tree exists
        if self.report_tree is not None:
            self.generate_report()
    
    def create_admin_dashboard_tab(self):
        filter_frame = ttk.LabelFrame(self.admin_dashboard_frame, text="Report Generation & Filters")
        filter_frame.pack(fill="x", padx=10, pady=10)
        
        # Date filter with radio buttons and calendar
        ttk.Label(filter_frame, text="Filter by Date:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        
        # Radio buttons for Date filter
        self.date_filter_var = tk.StringVar(value='All')
        date_all_radio = ttk.Radiobutton(filter_frame, text="All", variable=self.date_filter_var, 
                                        value='All', command=self.on_date_filter_change)
        date_all_radio.grid(row=0, column=1, padx=5, pady=5, sticky="w")
        
        date_specific_radio = ttk.Radiobutton(filter_frame, text="Specific Date:", variable=self.date_filter_var, 
                                             value='Specific', command=self.on_date_filter_change)
        date_specific_radio.grid(row=0, column=2, padx=5, pady=5, sticky="w")
        
        # Calendar for date selection
        self.date_calendar = DateEntry(filter_frame, width=15, background='darkblue',
                                      foreground='white', borderwidth=2, 
                                      date_pattern='dd-mm-yyyy', state='disabled')
        self.date_calendar.grid(row=0, column=3, padx=5, pady=5)
        self.date_calendar.bind('<<DateEntrySelected>>', lambda e: self.generate_report() if self.report_tree else None)
        
        # Service Category filter
        ttk.Label(filter_frame, text="Filter by Service Category:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.service_category_filter_var = tk.StringVar(value='All')
        category_filter_options = ['All', 'Sports', 'Non-Sports']
        self.category_filter_menu = ttk.Combobox(filter_frame, textvariable=self.service_category_filter_var, 
                                          values=category_filter_options, state='readonly')
        self.category_filter_menu.set('All')
        self.category_filter_menu.grid(row=1, column=1, columnspan=2, padx=5, pady=5, sticky="ew")
        self.category_filter_menu.bind('<<ComboboxSelected>>', self.update_service_filter_dropdown)
        
        # Service filter (Dynamic based on service category)
        ttk.Label(filter_frame, text="Filter by Service:").grid(row=2, column=0, padx=5, pady=5, sticky="w")
        self.service_var_filter = tk.StringVar(value='All')
        self.service_filter_menu = ttk.Combobox(filter_frame, textvariable=self.service_var_filter, 
                                   state='readonly')
        self.service_filter_menu.set('All')
        self.service_filter_menu.grid(row=2, column=1, columnspan=2, padx=5, pady=5, sticky="ew")
        self.service_filter_menu.bind('<<ComboboxSelected>>', lambda e: self.generate_report() if self.report_tree else None)
        
        # Category filter (Railway/Non Railway)
        ttk.Label(filter_frame, text="Filter by Category:").grid(row=3, column=0, padx=5, pady=5, sticky="w")
        self.category_var_filter = tk.StringVar(value='All')
        category_options = ['All', 'Railway', 'Non Railway']
        category_menu = ttk.Combobox(filter_frame, textvariable=self.category_var_filter, 
                                   values=category_options, state='readonly')
        category_menu.set('All')
        category_menu.grid(row=3, column=1, columnspan=2, padx=5, pady=5, sticky="ew")
        category_menu.bind('<<ComboboxSelected>>', lambda e: self.generate_report() if self.report_tree else None)
        
        # ADDED: Filter control buttons frame
        button_frame = ttk.Frame(filter_frame)
        button_frame.grid(row=4, column=0, columnspan=4, padx=5, pady=10)
        
        # Apply Filters button
        generate_button = ttk.Button(button_frame, text="Apply Filters", command=self.generate_report)
        generate_button.pack(side="left", padx=5)
        
        # ADDED: Clear Filter button
        clear_button = ttk.Button(button_frame, text="Clear Filter", command=self.clear_all_filters, 
                                 style='Accent.TButton')
        clear_button.pack(side="left", padx=5)
        
        results_frame = ttk.LabelFrame(self.admin_dashboard_frame, text="All Services Data (Auto-saved to Excel)")
        results_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        columns = ('Token No', 'Date', 'Name', 'Service Category', 'Service', 'Category', 'Fee')
        self.report_tree = ttk.Treeview(results_frame, columns=columns, show='headings')
        
        # Configure Treeview with service category-based color contrast - FIXED: Same size for bold
        self.report_tree.tag_configure('sports', background='#ADD8E6')  # Light Blue for sports
        self.report_tree.tag_configure('non_sports', background='#90EE90')  # Light Green for non-sports
        # FIXED: Railway bold styling for treeview - SAME SIZE
        self.report_tree.tag_configure('sports_railway', background='#ADD8E6', font=('TkDefaultFont', 10, 'bold'))
        
        for col in columns:
            self.report_tree.heading(col, text=col)
            if col == 'Token No':
                self.report_tree.column(col, width=80)
            elif col in ['Service Category', 'Service']:
                self.report_tree.column(col, width=110)
            else:
                self.report_tree.column(col, width=100)
        self.report_tree.pack(fill="both", expand=True)
        
        # Initialize service filter dropdown NOW that report_tree exists
        self.update_service_filter_dropdown()
        
        # Total amount display frame
        total_frame = ttk.Frame(results_frame)
        total_frame.pack(side="bottom", fill="x", padx=5, pady=5)
        
        self.total_amount_var = tk.StringVar()
        total_label = ttk.Label(total_frame, textvariable=self.total_amount_var, 
                               font=("Arial", 12, "bold"), foreground="green")
        total_label.pack(side="right")
        
        export_frame = ttk.LabelFrame(self.admin_dashboard_frame, text="Export Filtered Data")
        export_frame.pack(fill="x", padx=10, pady=10)
        
        # Export buttons
        excel_button = ttk.Button(export_frame, text="Download as Excel", command=self.export_filtered_excel)
        excel_button.pack(side="left", padx=10, pady=5)
        
        pdf_button = ttk.Button(export_frame, text="Download as PDF", command=self.export_filtered_pdf)
        pdf_button.pack(side="left", padx=10, pady=5)
        
        # Show data info with color legend
        info_frame = ttk.Frame(export_frame)
        info_frame.pack(side="right", padx=10, pady=5)
        
        info_label = ttk.Label(info_frame, text="All data auto-saved to park_services_dashboard.xlsx", 
                              font=("Arial", 9), foreground="green")
        info_label.pack()
        
        # Updated color legend
        legend_label = ttk.Label(info_frame, text="🔵 Sports  🟢 Non-Sports  📝 Railway=Bold", 
                                font=("Arial", 8), foreground="gray")
        legend_label.pack()
    
    def generate_report(self):
        """Generate report only if report_tree exists with Railway bold formatting"""
        if self.report_tree is None:
            return
            
        for item in self.report_tree.get_children():
            self.report_tree.delete(item)
        
        # Get filter values
        date_filter_type = self.date_filter_var.get()
        selected_date = self.date_calendar.get() if date_filter_type == 'Specific' else 'All'
        selected_service_category = self.service_category_filter_var.get()
        selected_service = self.service_var_filter.get()
        selected_category = self.category_var_filter.get()
        
        total_amount = 0
        filtered_count = 0
        
        for record in self.all_data:
            # Date filter
            if date_filter_type == 'All':
                date_match = True
            else:
                date_match = (record['Date'] == selected_date)
            
            # Service category filter with safe access
            service_category = record.get('Service Category', self.get_service_category(record.get('Service', 'Cricket')))
            service_category_match = (selected_service_category == 'All' or service_category == selected_service_category)
            
            service_match = (selected_service == 'All' or record.get('Service', '') == selected_service)
            category_match = (selected_category == 'All' or record.get('Category', '') == selected_category)
            
            if date_match and service_category_match and service_match and category_match:
                # Ensure Token No is displayed with 4 digits
                display_token_no = f"{int(record['Token No']):04d}" if str(record['Token No']).isdigit() else record['Token No']
                
                # Determine tag based on service category and Railway status for color coding
                category = record.get('Category', '')
                if service_category == 'Sports' and category == 'Railway':
                    tag = 'sports_railway'  # Bold Railway sports
                elif service_category == 'Sports':
                    tag = 'sports'
                else:
                    tag = 'non_sports'
                
                self.report_tree.insert("", "end", 
                                      values=(display_token_no, record.get('Date', ''), record.get('Name', ''), 
                                            service_category, record.get('Service', ''), 
                                            record.get('Category', ''), record.get('Fee', 0)),
                                      tags=(tag,))  # Apply color tag
                
                total_amount += record.get('Fee', 0)
                filtered_count += 1
        
        # Update total amount display
        if hasattr(self, 'total_amount_var'):
            self.total_amount_var.set(f"Total Collection: Rs. {total_amount:,} ({filtered_count} tokens)")
    
    def get_filtered_dataframe(self):
        """Get filtered data based on current filter settings"""
        date_filter_type = self.date_filter_var.get()
        selected_date = self.date_calendar.get() if date_filter_type == 'Specific' else 'All'
        selected_service_category = self.service_category_filter_var.get()
        selected_service = self.service_var_filter.get()
        selected_category = self.category_var_filter.get()
        
        filtered_data = []
        for record in self.all_data:
            if date_filter_type == 'All':
                date_match = True
            else:
                date_match = (record.get('Date', '') == selected_date)
            
            service_category = record.get('Service Category', self.get_service_category(record.get('Service', 'Cricket')))
            service_category_match = (selected_service_category == 'All' or service_category == selected_service_category)
            service_match = (selected_service == 'All' or record.get('Service', '') == selected_service)
            category_match = (selected_category == 'All' or record.get('Category', '') == selected_category)
            
            if date_match and service_category_match and service_match and category_match:
                # Ensure Token No is formatted with leading zeros
                record_copy = record.copy()
                record_copy['Token No'] = f"{int(record['Token No']):04d}" if str(record['Token No']).isdigit() else record['Token No']
                # Ensure Service Category is present
                record_copy['Service Category'] = service_category
                filtered_data.append(record_copy)
        
        return pd.DataFrame(filtered_data)
    
    def export_filtered_excel(self):
        """Export filtered data to Excel with professional styling"""
        try:
            df = self.get_filtered_dataframe()
            
            if df.empty:
                messagebox.showwarning("No Data", "No data matches the current filter.")
                return
            
            # Create filename
            date_filter_type = self.date_filter_var.get()
            selected_date = self.date_calendar.get() if date_filter_type == 'Specific' else 'All_Dates'
            selected_service_category = self.service_category_filter_var.get()
            selected_service = self.service_var_filter.get()
            selected_category = self.category_var_filter.get()
            
            filter_name = f"{selected_date}_{selected_service_category}_{selected_service}_{selected_category}".replace(" ", "_").replace("/", "-")
            filename = f"Park_Services_Report_{filter_name}.xlsx"
            
            # Save with professional styling using openpyxl
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Park Services Report')
                self.style_excel_professional(writer.book.active, df)
            
            messagebox.showinfo("Success", f"📊 Park Services Excel report exported!\n\n📄 File: {filename}\n🎫 Tokens: {len(df)}\n💰 Total Collection: Rs. {df['Fee'].sum():,}\n\n🎨 Features:\n• Blue: Sports Services\n• Green: Non-Sports Services\n• Bold: Railway Categories in Sports")
            
        except Exception as e:
            messagebox.showerror("Export Error", f"An error occurred: {str(e)}")
    
    def export_filtered_pdf(self):
        """Export filtered data to PDF with total amount and service category-based design"""
        try:
            df = self.get_filtered_dataframe()
            
            if df.empty:
                messagebox.showwarning("No Data", "No data matches the current filter.")
                return
            
            # Create filename and title
            date_filter_type = self.date_filter_var.get()
            selected_date = self.date_calendar.get() if date_filter_type == 'Specific' else 'All_Dates'
            selected_service_category = self.service_category_filter_var.get()
            selected_service = self.service_var_filter.get()
            selected_category = self.category_var_filter.get()
            
            filter_name = f"{selected_date}_{selected_service_category}_{selected_service}_{selected_category}".replace(" ", "_").replace("/", "-")
            filename = f"Park_Services_Report_{filter_name}.pdf"
            
            # Create title for PDF
            title_parts = []
            if selected_date != 'All_Dates':
                title_parts.append(f"Date: {selected_date}")
            if selected_service_category != 'All':
                title_parts.append(f"Service Category: {selected_service_category}")
            if selected_service != 'All':
                title_parts.append(f"Service: {selected_service}")
            if selected_category != 'All':
                title_parts.append(f"Category: {selected_category}")
            
            title = "Park Services Collection Report"
            if title_parts:
                title += f" - {' | '.join(title_parts)}"
            
            # Export to PDF with total row
            if self.dataframe_to_pdf(df, filename, title):
                total_amount = df['Fee'].sum()
                messagebox.showinfo("Success", f"📄 Park Services PDF exported!\n\n📄 File: {filename}\n🎫 Tokens: {len(df)}\n💰 Collection: Rs. {total_amount:,}\n\n✨ Features:\n• Service category-based colors\n• Railway categories in bold\n• Total collection included")
            else:
                messagebox.showerror("Error", "Failed to create PDF report.")
                
        except Exception as e:
            messagebox.showerror("Export Error", f"An error occurred: {str(e)}")

if __name__ == "__main__":
    try:
        import pandas as pd
        from tkcalendar import DateEntry
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_pdf import PdfPages
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import time
    except ImportError:
        print("Please install required libraries: pip install pandas openpyxl tkcalendar matplotlib")
        exit()
    
    root = tk.Tk()
    app = ParkServicesTrackerApp(root)
    root.mainloop()
